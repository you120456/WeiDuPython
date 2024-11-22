from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# 加载Excel文件
wb = load_workbook('example.xlsx')

# 设置新的背景颜色和字体样式
fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # 白色背景
font = Font(name='Arial', size=12, bold=True, color='000000')  # 黑色字体，Arial，12号字，加粗

# 遍历所有sheet
for sheet in wb:
    # 设置sheet的背景颜色和字体样式
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            cell.fill = fill
            cell.font = font

# 保存修改后的Excel文件
wb.save('example_modified.xlsx')

# 保存修改后的Excel文件
wb.save('D:/util-work/PythonProject/PythonProject/report/2024-07-31巴基斯坦自动日报.xlsx')