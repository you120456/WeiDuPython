import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def beautify_excel(file_path):
    # 获取需美化的Excel文件，并逐一打开工作簿，工作表
    wb = openpyxl.load_workbook(file_path)
    # 获取当前活跃的sheet页,默认就是第一个sheet页
    # ws = wb.active

    # 获取所有sheet的名称
    # ws = wb.sheetnames

    # 获取所有sheet的名称
    # sheet_names = wb.sheetnames

    print(wb.sheetnames)


    sheet_names =wb.sheetnames


    # 遍历所有的sheet
    for sheet_name in sheet_names:
        # 获取工作表的尺寸，用于循环设置每个单元格的样式
        print(sheet_name)
        sheet=wb[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column

        print('sheet页名称：')
        print(sheet)
        print('最大行数：')
        print(max_row)
        print('最大列数：')
        print(+max_col)



    # 设置字体样式
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    cell_font = Font(name='Arial')

    # 设置对齐方式
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 设置填充颜色
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    # 设置边框样式
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


    # 确认数据的范围
    # max_row = ws.max_row
    # max_col = ws.max_column



    # 应用样式到标题行
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # 应用样式到数据单元格
    for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = cell_font
            cell.alignment = center_alignment
            cell.border = thin_border

    # 自动调整列宽，包括表头行
    for col in sheet.iter_cols(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # 设置行高
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        sheet.row_dimensions[row[0].row].height = 20  # 设置行高为20

    # 保存工作簿
    wb.save(file_path)
    print(f"{file_path} 已成功美化")

# 调用函数
# beautify_excel("your_excel_file.xlsx")
