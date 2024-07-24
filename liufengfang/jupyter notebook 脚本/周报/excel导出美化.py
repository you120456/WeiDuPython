import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def beautify_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 设置字体样式
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    cell_font = Font(name='Arial')

    # 设置对齐方式
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 设置填充颜色
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    # 设置边框样式
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 确认数据的范围
    max_row = ws.max_row
    max_col = ws.max_column

    # 应用样式到标题行
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # 应用样式到数据单元格
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = cell_font
            cell.alignment = center_alignment
            cell.border = thin_border

    # 自动调整列宽，包括表头行
    for col in ws.iter_cols(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 设置行高
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        ws.row_dimensions[row[0].row].height = 20  # 设置行高为20

    # 保存工作簿
    wb.save(file_path)
    print(f"{file_path} 已成功美化")

# 调用函数
# beautify_excel("your_excel_file.xlsx")
