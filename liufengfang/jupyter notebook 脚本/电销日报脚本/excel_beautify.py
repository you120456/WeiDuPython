import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# import shutil

def beautify_excel(file_path):
    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"无法加载工作簿: {e}")
        return

    # 设置字体样式
    header_font = Font(name='Arial', bold=True, color='FFFFFF')
    cell_font = Font(name='Arial')

    # 设置对齐方式
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 设置填充颜色
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    # 设置边框样式
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 遍历所有工作表
    for ws in wb.worksheets:
        # 确认数据的范围
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row == 0 or max_col == 0:
            continue

        # 应用样式到标题行
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border

        # 应用样式到数据单元格
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.font = cell_font
                cell.alignment = center_alignment
                cell.border = thin_border

        # 自动调整列宽，包括表头行
        for col in ws.iter_cols(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        # 计算中文字符的宽度，假设每个中文字符宽度为2个字符的宽度
                        cell_value = str(cell.value)
                        if any('\u4e00' <= char <= '\u9fff' for char in cell_value):
                            cell_length = len(cell_value) * 2  # 中文字符宽度估算
                        else:
                            cell_length = len(cell_value)
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception as e:
                    print(f"处理单元格 {cell.coordinate} 时出错: {e}")

            # 计算列宽，使用更大的缓冲量
            estimated_width = max_length * 1.5  # 假设每个字符的实际宽度为1.5个字符的宽度
            adjusted_width = max(estimated_width, 25)  # 设置最小宽度为25

            ws.column_dimensions[column].width = adjusted_width

        # 设置行高
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            ws.row_dimensions[row[0].row].height = 20  # 设置行高为20

    # 保存工作簿前备份原文件
    # backup_file_path = file_path + '.bak'
    # shutil.copy(file_path, backup_file_path)
    # print(f"原文件已备份至: {backup_file_path}")

    # 保存工作簿
    try:
        wb.save(file_path)
        print(f"{file_path} 已成功美化")
    except Exception as e:
        print(f"保存工作簿时出错: {e}")

# 调用函数
# beautify_excel("your_excel_file.xlsx")
